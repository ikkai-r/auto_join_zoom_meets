import openpyxl as op
import webbrowser as wb

import schedule as sc
import time as t

# get the zoom meetings
dataframe = op.load_workbook("zoom_meets.xlsx")
df1 = dataframe.active


def process_meeting():
    # read the cols and rows
    col_count = 0

    # lists for the zoom meetings per day
    mon_meets = []
    tues_meets = []
    thurs_meets = []
    fri_meets = []

    for row in range(1, df1.max_row):
        zoom_meet = ''
        zoom_time = ''
        for col in df1.iter_cols(1, df1.max_column):

            # get zoom link for the row
            if col_count == 1:
                zoom_meet = col[row].value
            elif col_count == 2:
                zoom_time = col[row].value

            # start from 0 if next row
            if col_count < 3:
                col_count += 1
            else:
                # check what day the zoom meet occurs
                meet_days = str(col[row].value).split()

                for mdays in meet_days:
                    if mdays == 'M':
                        mon_meets.append(zoom_meet + '|' + str(zoom_time))

                    if mdays == 'T':
                        tues_meets.append(zoom_meet + '|' + str(zoom_time))

                    if mdays == 'Th':
                        thurs_meets.append(zoom_meet + '|' + str(zoom_time))

                    if mdays == 'F':
                        fri_meets.append(zoom_meet + '|' + str(zoom_time))

                # reset col_count
                col_count = 0

    week_meets = [mon_meets, tues_meets, thurs_meets, fri_meets]
    join_meets(week_meets)


def join_meets(week_meets):
    # process monday meets
    for mon_meet in week_meets[0]:
        mon_meet_str = mon_meet.split("|")
        meet_time = mon_meet_str[1]
        meet_link = mon_meet_str[0]
        sc.every().monday.at(meet_time).do(open_meet, meet_link)

    # process tuesday meets
    for tues_meet in week_meets[1]:
        tues_meet_str = tues_meet.split("|")
        meet_time = tues_meet_str[1]
        meet_link = tues_meet_str[0]
        sc.every().tuesday.at(meet_time).do(open_meet, meet_link)

    for thurs_meet in week_meets[2]:
        thurs_meet_str = thurs_meet.split("|")
        meet_time = thurs_meet_str[1]
        meet_link = thurs_meet_str[0]
        sc.every().thursday.at(meet_time).do(open_meet, meet_link)

    for fri_meet in week_meets[3]:
        fri_meet_str = fri_meet.split("|")
        meet_time = fri_meet_str[1]
        meet_link = fri_meet_str[0]
        sc.every().friday.at(meet_time).do(open_meet, meet_link)


def open_meet(meet_link):
    wb.open(meet_link)


def main():
    process_meeting()
    while True:
        sc.run_pending()
        t.sleep(1)


if __name__ == "__main__":
    print('running...')
    main()
